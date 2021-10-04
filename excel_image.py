import cv2          #导入OpenCV库
import xlsxwriter   #利用这个调整行高列宽
import openpyxl     #利用这个填充颜色
import numpy as np  #数据存储的方式，用此种方式处理数据比列表高效
import pandas as pd #数据存储的方式，用此种方式处理数据比列表高效

class ImageToExcel():
    #初始化
    def __init__(self,image_path,excel_path):
        self.imgviewx = cv2.imread(image_path,cv2.IMREAD_COLOR)
        self.excel_path = excel_path
    
    # excel行高列宽调整
    def excel_size(self):
        workbook = xlsxwriter.Workbook(self.excel_path)
        worksheet = workbook.add_worksheet('test')
        worksheet.set_column('A:CAA', 1)
        for x in range(500): 
            worksheet.set_row(x, 8.4)
        workbook.close()
    
    #rgb转16进制颜色码
    def ten2_16(self,num):
        tmp = hex(num).replace('0x', '')
        if len(tmp) > 1:
            return tmp
        else:
            '0' + tmp
    
    #获取像素数据并转化为16进制
    def get_rgb_data(self):
        self.excel_size()
        r_array = np.array(self.imgviewx)[:,:,2]
        r = pd.DataFrame(r_array)
        data_r = r.applymap(self.ten2_16)
        g_array = np.array(self.imgviewx)[:,:,1]
        g = pd.DataFrame(g_array)
        data_g = g.applymap(self.ten2_16)
        b_array = np.array(self.imgviewx)[:,:,0]
        b = pd.DataFrame(b_array)
        data_b = b.applymap(self.ten2_16)
        return (data_r+data_g+data_b).values
    
    #颜色填充
    def color_fill(self):
        rgb_list = self.get_rgb_data()
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb['test']
        print("正在填充，请稍候...")
        for x,tmp1 in list(enumerate(rgb_list)):
            for y ,tmp2 in list(enumerate(tmp1)):
                ws.cell(x+1,y+1).fill = \
                openpyxl.styles.fills.GradientFill(stop=[str(tmp2),str(tmp2)])
        wb.save(self.excel_path)
        print("填充完成！")

excel_path = 'export.xlsx'
image_path = 'import.png'
image = ImageToExcel(image_path,excel_path)
image.color_fill()